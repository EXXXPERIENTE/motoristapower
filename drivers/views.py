import logging
import io
import threading
from datetime import datetime, date

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.db.models import Count, Q, Sum, Avg
from django.db import IntegrityError
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .forms import MotoristaForm
from .models import Motorista

# Configuração de logger
logger = logging.getLogger(__name__)


# ✅ NOVAS VIEWS PÚBLICAS (SEM LOGIN)

def pagina_inicial(request):
    """Página inicial pública - qualquer um acessa sem login"""
    return render(request, 'drivers/pagina_inicial.html')


def pagina_sucesso(request):
    """Página de sucesso após cadastro - pública"""
    return render(request, 'drivers/sucesso.html')


def cadastro_motorista(request):
    """
    ✅ CADASTRO PÚBLICO - qualquer pessoa pode se cadastrar sem login
    """
    if request.method == 'POST':
        form = MotoristaForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                motorista = form.save(commit=False)

                # ✅ SE o usuário ESTIVER LOGADO, associa a ele
                if request.user.is_authenticated:
                    motorista.user = request.user
                # ✅ SE for USUÁRIO ANÔNIMO, cria sem usuário (pode fazer login depois)
                else:
                    motorista.user = None

                motorista.save()

                messages.success(request, f"✅ Motorista {motorista.nome_completo} cadastrado com sucesso!")

                # ✅ Redireciona para página de sucesso (pública)
                return redirect('drivers:sucesso')

            except IntegrityError:
                messages.error(request, '❌ Já existe um motorista com este CPF ou CNH!')
            except Exception as e:
                messages.error(request, f'❌ Erro ao cadastrar: {e}')
        else:
            messages.error(request, '❌ Erro no formulário. Verifique os dados.')
    else:
        form = MotoristaForm()

    return render(request, 'drivers/cadastro_motorista.html', {'form': form, 'editando': False})


# 🔐 VIEWS PRIVADAS (COM LOGIN REQUIRED)

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'drivers/dashboard.html'
    login_url = '/accounts/login/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Se não for Staff, ele não deveria ver esta view, mas o LoginRequiredMixin já protege o acesso.
        # Motoristas comuns (não staff) que logarem, verão o dashboard, mas com dados limitados se for o caso.

        total_motoristas = Motorista.objects.count()
        motoristas_ativos = Motorista.objects.filter(status='ATIVO').count()
        motoristas_inativos = Motorista.objects.filter(status='INATIVO').count()
        estados_stats = Motorista.objects.values('estado').annotate(
            total=Count('id')
        ).order_by('-total')[:5]
        ultimos_cadastros = Motorista.objects.all().order_by('-created_at')[:5]

        context.update({
            'total_motoristas': total_motoristas,
            'motoristas_ativos': motoristas_ativos,
            'motoristas_inativos': motoristas_inativos,
            'estados_stats': estados_stats,
            'ultimos_cadastros': ultimos_cadastros,
        })

        return context


class MotoristaListView(LoginRequiredMixin, ListView):
    model = Motorista
    template_name = 'drivers/motorista_list.html'
    context_object_name = 'motoristas'
    paginate_by = 10
    ordering = ['-created_at']
    login_url = '/accounts/login/'

    def get_queryset(self):
        # 🚨 FLUXO IDEAL DO MOTORISTA COMUM: Só pode ver o próprio cadastro
        if not self.request.user.is_staff:
            # Tenta filtrar pelo usuário logado.
            try:
                return Motorista.objects.filter(user=self.request.user)
            except Motorista.DoesNotExist:
                return Motorista.objects.none()

        # Admin/Staff veem todos, aplicando filtros de busca
        queryset = super().get_queryset()
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome_completo__icontains=search) |
                Q(cpf__icontains=search) |
                Q(cnh_numero__icontains=search) |
                Q(cidade__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # O Superusuário (is_staff=True) vê o total geral.
        if self.request.user.is_staff:
            context['total_motoristas'] = Motorista.objects.count()
            context['motoristas_ativos'] = Motorista.objects.filter(status='ATIVO').count()
            context['motoristas_inativos'] = Motorista.objects.filter(status='INATIVO').count()
        else:
            # O Motorista Comum (is_staff=False) vê apenas seu registro (0 ou 1)
            context['total_motoristas'] = self.get_queryset().count()
            context['motoristas_ativos'] = self.get_queryset().filter(status='ATIVO').count()
            context['motoristas_inativos'] = self.get_queryset().filter(status='INATIVO').count()

        context['current_status'] = self.request.GET.get('status', '')
        context['current_search'] = self.request.GET.get('search', '')
        return context


class MotoristaUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Motorista
    form_class = MotoristaForm
    template_name = 'drivers/cadastro_motorista.html'
    success_url = reverse_lazy('drivers:motorista_list')
    login_url = '/accounts/login/'

    def test_func(self):
        motorista = self.get_object()
        # Permite se for Staff/Admin OU se o usuário logado for o dono do objeto
        return self.request.user.is_staff or motorista.user == self.request.user

    def handle_no_permission(self):
        messages.error(self.request, "Acesso negado. Você só pode editar seu próprio cadastro.")
        return redirect('drivers:dashboard')

    def get_success_url(self):
        if self.request.user.is_staff:
            return reverse_lazy('drivers:motorista_list')
        return reverse_lazy('drivers:dashboard')

    def form_valid(self, form):
        motorista = form.save(commit=False)
        if not motorista.user:
            motorista.user = self.request.user
        motorista.save()

        messages.success(self.request, 'Motorista atualizado com sucesso!')
        return redirect(self.get_success_url())


class MotoristaDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Motorista
    template_name = 'drivers/motorista_confirm_delete.html'
    success_url = reverse_lazy('drivers:motorista_list')
    login_url = '/accounts/login/'

    def test_func(self):
        # Apenas Staff/Admin pode deletar
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, "Acesso negado. Você não tem permissão para deletar cadastros.")
        return redirect('drivers:dashboard')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Motorista excluído com sucesso!')
        return super().delete(request, *args, **kwargs)


@login_required
def lista_motoristas(request):
    if not request.user.is_staff:
        motoristas = Motorista.objects.filter(user=request.user)
    else:
        motoristas = Motorista.objects.all()

    contexto = {
        'motoristas': motoristas,
        'titulo_pagina': 'Lista de Motoristas'
    }
    return render(request, 'drivers/lista_motoristas.html', contexto)


@login_required
def relatorio_estatisticas(request):
    if not request.user.is_staff:
        messages.error(request, "Acesso negado. Apenas administradores podem ver as estatísticas.")
        return redirect('drivers:dashboard')

    total_motoristas = Motorista.objects.count()
    status_stats = Motorista.objects.values('status').annotate(total=Count('id')).order_by('-total')
    estado_stats = Motorista.objects.values('estado').annotate(total=Count('id')).order_by('-total')
    categoria_stats = Motorista.objects.values('cnh_categoria').annotate(total=Count('id')).order_by('-total')
    total_salarios = Motorista.objects.filter(salario__isnull=False).aggregate(total=Sum('salario'))['total'] or 0
    salario_medio = Motorista.objects.filter(salario__isnull=False).aggregate(medio=Avg('salario'))['medio'] or 0

    idade_media = None
    if total_motoristas > 0:
        idades = [motorista.idade for motorista in Motorista.objects.all() if motorista.data_nascimento]
        if idades:
            idade_media = sum(idades) / len(idades)

    if salario_medio:
        salario_medio = f'{salario_medio:,.2f}'
    if total_salarios:
        total_salarios = f'{total_salarios:,.2f}'

    context = {
        'total_motoristas': total_motoristas,
        'status_stats': status_stats,
        'estado_stats': estado_stats,
        'categoria_stats': categoria_stats,
        'total_salarios': total_salarios,
        'salario_medio': salario_medio,
        'idade_media': f'{idade_media:.2f}' if idade_media else None,
    }

    return render(request, 'drivers/relatorio_estatisticas.html', context)


@login_required
def relatorio_excel(request):
    if not request.user.is_staff:
        messages.error(request, "Acesso negado. Apenas administradores podem gerar relatórios.")
        return redirect('drivers:dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Motoristas"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:L1')
    ws['A1'] = "RELATÓRIO DE MOTORISTAS - MOTORISTAPOWER"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:L2')
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].alignment = center_align

    headers = ['ID', 'Nome', 'CPF', 'Data Nasc.', 'Idade', 'Telefone', 'Cidade/UF',
               'CNH', 'Categoria', 'Status', 'Salário', 'Data Cadastro']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    motoristas = Motorista.objects.all().order_by('nome_completo')
    row = 4
    for row_num, motorista in enumerate(motoristas, 5):
        ws.cell(row=row_num, column=1, value=motorista.id)
        ws.cell(row=row_num, column=2, value=motorista.nome_completo or 'NÃO INFORMADO')
        ws.cell(row=row_num, column=3, value=motorista.cpf_formatado)
        ws.cell(row=row_num, column=4,
                value=motorista.data_nascimento.strftime('%d/%m/%Y') if motorista.data_nascimento else '')
        ws.cell(row=row_num, column=5, value=motorista.idade)
        ws.cell(row=row_num, column=6, value=motorista.telefone or '')
        ws.cell(row=row_num, column=7, value=f"{motorista.cidade or ''}/{motorista.estado or ''}")
        ws.cell(row=row_num, column=8, value=motorista.cnh_numero or '')
        ws.cell(row=row_num, column=9, value=motorista.cnh_categoria or '')
        ws.cell(row=row_num, column=10, value=motorista.get_status_display())
        ws.cell(row=row_num, column=11, value=float(motorista.salario) if motorista.salario else 0)
        ws.cell(row=row_num, column=12, value=motorista.created_at.strftime('%d/%m/%Y %H:%M'))
        row = row_num

    column_widths = [8, 30, 15, 12, 8, 15, 15, 15, 10, 12, 12, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.merge_cells(f'A{row + 2}:L{row + 2}')
    ws.cell(row=row + 2, column=1, value=f"TOTAL DE MOTORISTAS: {motoristas.count()}")
    ws.cell(row=row + 2, column=1).font = Font(bold=True, color="366092")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_motoristas.xlsx"'

    wb.save(response)
    return response


@login_required
def relatorio_pdf(request):
    if not request.user.is_staff:
        messages.error(request, "Acesso negado. Apenas administradores podem gerar relatórios.")
        return redirect('drivers:dashboard')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Center',
        alignment=1,
        fontSize=14,
        spaceAfter=30
    ))

    elements = []

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,
        spaceAfter=30
    )

    elements.append(Paragraph("RELATÓRIO DE MOTORISTAS", title_style))
    elements.append(Paragraph(f"<b>MotoristaPower</b> - Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
                              styles['Center']))
    elements.append(Spacer(1, 20))

    motoristas = Motorista.objects.all().order_by('nome_completo')

    if motoristas:
        data = [['ID', 'Nome', 'CPF', 'Idade', 'Cidade/UF', 'Status', 'CNH']]

        for motorista in enumerate(motoristas):
            data.append([
                str(motorista.id),
                motorista.nome_completo or 'NÃO INFORMADO',
                motorista.cpf_formatado,
                str(motorista.idade),
                f"{motorista.cidade or ''}/{motorista.estado or ''}",
                motorista.get_status_display(),
                motorista.cnh_categoria or 'N/I'
            ])

        table = Table(data, colWidths=[0.5 * inch, 2 * inch, 1.2 * inch, 0.6 * inch, 1 * inch, 0.8 * inch, 0.6 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (0, -1), 1, colors.black),
            ('GRID', (1, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph(f"<b>Total de Motoristas:</b> {motoristas.count()}", styles['Normal']))
        elements.append(Paragraph(f"<b>Ativos:</b> {motoristas.filter(status='ATIVO').count()}", styles['Normal']))
        elements.append(Paragraph(f"<b>Inativos:</b> {motoristas.filter(status='INATIVO').count()}", styles['Normal']))

    else:
        elements.append(Paragraph("Nenhum motorista cadastrado.", styles['Normal']))

    doc.build(elements)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_motoristas.pdf"'
    response.write(buffer.getvalue())
    buffer.close()

    return response


@login_required
def relatorio_estatisticas_excel(request):
    if not request.user.is_staff:
        messages.error(request, "Acesso negado. Apenas administradores podem gerar relatórios.")
        return redirect('drivers:dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estatísticas"

    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:E1')
    ws['A1'] = "RELATÓRIO ESTATÍSTICO - MOTORISTAPOWER"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].alignment = center_align

    total_motoristas = Motorista.objects.count()
    ativos = Motorista.objects.filter(status='ATIVO').count()
    inativos = Motorista.objects.filter(status='INATIVO').count()
    total_salarios = Motorista.objects.filter(salario__isnull=False).aggregate(total=Sum('salario'))['total'] or 0

    ws['A4'] = "ESTATÍSTICAS GERAIS"
    ws['A4'].font = Font(bold=True, size=12)

    total_salarios_formatado = f'R$ {total_salarios:,.2f}' if isinstance(total_salarios, (int, float)) else 'R$ 0,00'

    data_geral = [
        ['Total de Motoristas', total_motoristas],
        ['Motoristas Ativos', ativos],
        ['Motoristas Inativos', inativos],
        ['Folha de Pagamento Total', total_salarios_formatado],
    ]

    for row, (label, value) in enumerate(data_geral, 5):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = Font(bold=True)

    ws['A10'] = "DISTRIBUIÇÃO POR ESTADO"
    ws['A10'].font = Font(bold=True, size=12)

    estados = Motorista.objects.values('estado').annotate(total=Count('id')).order_by('-total')
    for row, estado in enumerate(estados, 11):
        ws.cell(row=row, column=1, value=estado['estado'] or 'NÃO INFORMADO')
        ws.cell(row=row, column=2, value=estado['total'])

    ws['D10'] = "DISTRIBUIÇÃO POR CATEGORIA CNH"
    ws['D10'].font = Font(bold=True, size=12)

    categorias = Motorista.objects.values('cnh_categoria').annotate(total=Count('id')).order_by('-total')
    for row, categoria in enumerate(categorias, 11):
        ws.cell(row=row, column=4, value=categoria['cnh_categoria'] or 'NÃO INFORMADA')
        ws.cell(row=row, column=5, value=categoria['total'])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estatisticas_motoristas.xlsx"'

    wb.save(response)
    return response